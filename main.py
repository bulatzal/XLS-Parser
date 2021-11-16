from openpyxl import load_workbook


def load():
    num = int(input('Введите номер книги: '))
    rows = sheet.max_row
    for i in range(2, rows):
        if sheet.cell(row=i, column=1).value == num:
            for j in range(1, 6):
                print(str(sheet.cell(row=1, column=j).value) + ' - ' + str(sheet.cell(row=i, column=j).value))


def create():
    rows = sheet.max_row
    columns = sheet.max_column
    sheet.cell(row=rows + 1, column=1).value = rows
    for i in range(2, columns + 1):
        book = input(str(sheet.cell(row=1, column=i).value) + ' - ')
        sheet.cell(row=rows + 1, column=i).value = book
    file.save('books.xlsx')


file = load_workbook('books.xlsx')
sheet = file['Лист1']
action = input('Выберите действие: /load - для просмотра, /create - для создания \n')
match action:
    case "/load":
        load()
    case "/create":
        create()
    case _:
        print('Такого действия нет')
